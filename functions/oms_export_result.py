import azure.functions as func
import csv
import pandas as pd
import json
from io import BytesIO, StringIO
import socket
import urllib.error
from azure.identity import DefaultAzureCredential
from databricks import sql
import smtplib
import email_config
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill, Font
import jwt
import base64
import smtplib
from cryptography.hazmat.primitives.ciphers import Cipher, algorithms, modes
from cryptography.hazmat.backends import default_backend
from cryptography.hazmat.primitives import padding

SECRET_KEY = "OMSTokenSecretKey@123".ljust(32, "x").encode()  # Đảm bảo đủ 32 bytes
IV = "abcd012345678910".encode()  # Đảm bảo đủ 16 bytes

# Email configuration
# gmail_password = email_config.email_password
# email_from = email_config.email_from
# host = email_config.smtp_host
# port = email_config.smtp_port

# email config local
gmail_password='phmbxqoetiaskrez'
host = 'smtp.gmail.com'
port='465'
email_from="bi-thailand@msc.masangroup.com"

# Databricks configuration
DATABRICKS_RESOURCE_ID = "2ff814a6-3304-4ab8-85cb-cd0e6f879c1d"
SERVER_HOSTNAME = "adb-1538821690907541.1.azuredatabricks.net"
HTTP_PATH = "/sql/1.0/warehouses/b6d556c5cf816ae6"
DATABASE = "default"

def main(req: func.HttpRequest) -> func.HttpResponse:
    headers = {
        "Access-Control-Allow-Origin": "*",
        "Access-Control-Allow-Methods": "OPTIONS,POST,GET",
        "Access-Control-Allow-Headers": "Content-Type,X-Amz-Date,Authorization,X-Api-Key,X-Amz-Security-Token"
    }
    
    if req.method == 'OPTIONS':
        return func.HttpResponse('', status_code=204, headers=headers)
    
    try:
        
        # Parse request JSON
        request_json = req.get_json()
        if not request_json:
            return func.HttpResponse(
                json.dumps({"error": "Missing or invalid JSON payload"}), 
                status_code=400, 
                headers=headers
            )

        # Lấy các tham số từ request
        year = request_json.get("Year")
        month = request_json.get("Month")
        factory_code = request_json.get("Factory_Code")
        pillar_code = request_json.get("Pillar_Code", None)
        is_eng = request_json.get("Is_ENG", 0)
        is_vie = request_json.get("Is_VIE", 1)
        email_to = request_json.get("Email")
        file_type = request_json.get("Type", "excel").lower()

        if not all([year, month, factory_code, email_to]):
            return func.HttpResponse(
                json.dumps({"error": "Missing required fields: Year, Month, Factory_Code, Email"}), 
                status_code=400, 
                headers=headers
            )

        # Databricks connection setup
        credential = DefaultAzureCredential()
        access_token = credential.get_token(DATABRICKS_RESOURCE_ID + "/.default").token

        with sql.connect(
            server_hostname=SERVER_HOSTNAME,
            http_path=HTTP_PATH,
            access_token=access_token,
            database=DATABASE
        ) as connection:
            with connection.cursor() as cursor:
                
                # Truy vấn Databricks với parameterized query
                query = f"""
                WITH SC AS (
                    SELECT S.YEAR, S.MONTH, S.FACTORY_CODE, P.ORDER_NO PILLAR_NO, LV.LEVEL_NO, SC.TYPE, SC.CODE,
                        CASE WHEN {is_eng} = 1 AND {is_vie} = 1 THEN CONCAT(SC.ENG_NAME, ' / ', SC.VIE_NAME)
                            WHEN {is_eng} = 1 AND {is_vie} = 0 THEN SC.ENG_NAME
                            WHEN {is_eng} = 0 AND {is_vie} = 1 THEN SC.VIE_NAME
                            ELSE '' END AS NAME,
                        SUM(CASE WHEN S.SCORE <> -1 THEN S.SCORE END) AS TOTAL_SCORE,
                        COUNT(1) ACTION,
                        COUNT(CASE WHEN SCORE = -1 THEN 1 END) ACTION_NA,
                        ROUND(SUM(CASE WHEN S.SCORE <> -1 THEN S.SCORE END) / (COUNT(CASE WHEN SCORE <> -1 THEN 1 END) * 3)
                            * CASE WHEN P.PILLAR_CODE <> 'SAF' THEN 100 ELSE 1 END, 2) AS Assessment,
                        NULL Total_score_pillar, NULL Final_score, NULL Grade,
                        CASE WHEN C.CODE = 'SAF00' THEN
                            CASE WHEN SUM(CASE WHEN S.SCORE <> -1 THEN S.SCORE END) * 0.8 < 15
                                THEN ROUND(SUM(CASE WHEN S.SCORE <> -1 THEN S.SCORE END) * 0.8, 2)
                                ELSE 15 END
                        END AS REDUCT_SCORING,
                        CAST(ROUND((1 - (CASE WHEN C.CODE = 'SAF00' THEN
                            CASE WHEN SUM(CASE WHEN S.SCORE <> -1 THEN S.SCORE END) * 0.8 < 15
                                THEN SUM(CASE WHEN S.SCORE <> -1 THEN S.SCORE END) * 0.8
                                ELSE 15 END
                        END / 15)) * 100, 0) AS STRING) || '%' AS COMPLIANCE,
                        C.CODE CHAPTER_CODE, C.PILLAR_CODE
                    FROM udp_wcm_dev.MFG_OMS.OMS_SCORE_F S
                    LEFT JOIN udp_wcm_dev.MFG_OMS.OMS_TYPE_D KPI ON S.ITEM_CODE = KPI.CODE AND KPI.TYPE = 'Item'
                    LEFT JOIN udp_wcm_dev.MFG_OMS.OMS_TYPE_D SC ON KPI.PARENT_CODE = SC.CODE AND SC.TYPE = 'Sub_chapter'
                    LEFT JOIN udp_wcm_dev.MFG_OMS.OMS_TYPE_D C ON SC.PARENT_CODE = C.CODE AND C.TYPE = 'Chapter'
                    LEFT JOIN udp_wcm_dev.MFG_OMS.OMS_PILLAR_D P ON P.PILLAR_CODE = C.PILLAR_CODE
                    LEFT JOIN udp_wcm_dev.MFG_OMS.OMS_LEVEL_TYPE_D LV ON LV.TYPE = SC.TYPE
                    WHERE S.YEAR = {year} AND S.MONTH = {month} AND S.FACTORY_CODE = '{factory_code}' AND P.PILLAR_CODE = '{pillar_code}'
                    GROUP BY S.YEAR, S.MONTH, S.FACTORY_CODE, P.ORDER_NO, LV.LEVEL_NO, SC.TYPE, SC.CODE, SC.ENG_NAME, SC.VIE_NAME,
                        C.CODE, C.PILLAR_CODE, P.PILLAR_CODE
                ),
                C AS (
                    SELECT S.YEAR, S.MONTH, S.FACTORY_CODE, P.ORDER_NO PILLAR_NO, LV.LEVEL_NO, C.TYPE, C.CODE,
                        CASE WHEN {is_eng} = 1 AND {is_vie} = 1 THEN CONCAT(C.ENG_NAME, ' / ', C.VIE_NAME)
                            WHEN {is_eng} = 1 AND {is_vie} = 0 THEN C.ENG_NAME
                            WHEN {is_eng} = 0 AND {is_vie} = 1 THEN C.VIE_NAME
                            ELSE '' END AS NAME,
                        SUM(CASE WHEN S.SCORE <> -1 THEN S.SCORE END) AS TOTAL_SCORE,
                        COUNT(1) AS ACTION,
                        COUNT(CASE WHEN SCORE = -1 THEN 1 END) AS ACTION_NA,
                        ROUND(SUM(CASE WHEN S.SCORE <> -1 THEN S.SCORE END) / (COUNT(CASE WHEN SCORE <> -1 THEN 1 END) * 3)
                            * CASE WHEN P.PILLAR_CODE <> 'SAF' THEN 100 ELSE 1 END, 2) AS ASSESSMENT,
                        C.PILLAR_CODE
                    FROM udp_wcm_dev.MFG_OMS.OMS_SCORE_F S
                    LEFT JOIN udp_wcm_dev.MFG_OMS.OMS_TYPE_D KPI ON S.ITEM_CODE = KPI.CODE AND KPI.TYPE = 'Item'
                    LEFT JOIN udp_wcm_dev.MFG_OMS.OMS_TYPE_D SC ON KPI.PARENT_CODE = SC.CODE AND SC.TYPE = 'Sub_chapter'
                    LEFT JOIN udp_wcm_dev.MFG_OMS.OMS_TYPE_D C ON SC.PARENT_CODE = C.CODE AND C.TYPE = 'Chapter'
                    LEFT JOIN udp_wcm_dev.MFG_OMS.OMS_PILLAR_D P ON P.PILLAR_CODE = C.PILLAR_CODE
                    LEFT JOIN udp_wcm_dev.MFG_OMS.OMS_LEVEL_TYPE_D LV ON LV.TYPE = C.TYPE
                    WHERE S.YEAR = {year} AND S.MONTH = {month} AND S.FACTORY_CODE = '{factory_code}' AND P.PILLAR_CODE = '{pillar_code}'
                        AND C.CODE <> 'SAF00'
                    GROUP BY S.YEAR, S.MONTH, S.FACTORY_CODE, P.ORDER_NO, LV.LEVEL_NO, C.TYPE, C.CODE, C.ENG_NAME, C.VIE_NAME,
                        C.CODE, C.PILLAR_CODE, P.PILLAR_CODE
                ),
                C_SAF00 AS (
                    SELECT C.YEAR, C.MONTH, C.FACTORY_CODE, C.PILLAR_NO, C.LEVEL_NO, T.TYPE, T.CODE,
                        CASE WHEN {is_eng} = 1 AND {is_vie} = 1 THEN CONCAT(T.ENG_NAME, ' / ', T.VIE_NAME)
                            WHEN {is_eng} = 1 AND {is_vie} = 0 THEN T.ENG_NAME
                            WHEN {is_eng} = 0 AND {is_vie} = 1 THEN T.VIE_NAME
                            ELSE '' END AS NAME,
                        SUM(C.TOTAL_SCORE) TOTAL_SCORE,
                        SUM(C.ACTION) ACTION,
                        SUM(C.ACTION_NA) ACTION_NA,
                        ROUND(SUM(C.ASSESSMENT), 2) ASSESSMENT,
                        ROUND(SC.REDUCT_SCORING, 2) REDUCT_SCORING,
                        CAST(ROUND(100 - (SC.REDUCT_SCORING / 15 * 8), 2) AS STRING) || '%' AS COMPLIANCE,
                        T.PILLAR_CODE
                    FROM C
                    INNER JOIN udp_wcm_dev.MFG_OMS.OMS_TYPE_D T ON T.TYPE = 'Chapter' AND T.CODE = 'SAF00'
                    LEFT JOIN (
                        SELECT SC.CHAPTER_CODE, SUM(SC.REDUCT_SCORING) REDUCT_SCORING
                        FROM SC
                        WHERE SC.CHAPTER_CODE = 'SAF00'
                        GROUP BY SC.CHAPTER_CODE
                    ) SC ON SC.CHAPTER_CODE = T.CODE
                    WHERE C.PILLAR_CODE = 'SAF'
                    GROUP BY C.YEAR, C.MONTH, C.FACTORY_CODE, C.PILLAR_NO, C.LEVEL_NO, T.TYPE, T.CODE,
                        CASE WHEN {is_eng} = 1 AND {is_vie} = 1 THEN CONCAT(T.ENG_NAME, ' / ', T.VIE_NAME)
                            WHEN {is_eng} = 1 AND {is_vie} = 0 THEN T.ENG_NAME
                            WHEN {is_eng} = 0 AND {is_vie} = 1 THEN T.VIE_NAME
                            ELSE '' END,
                        SC.REDUCT_SCORING, T.PILLAR_CODE
                )
                SELECT YEAR, MONTH, FACTORY_CODE, LEVEL_NO, TYPE, CODE, NAME, TOTAL_SCORE, ACTION, ACTION_NA, ASSESSMENT,
                    REDUCT_SCORING, COMPLIANCE,
                    TOTAL_SCORE_PILLAR, FINAL_SCORE, GRADE,
                    CREATED_BY, CREATED_DATE,
                    LV1 PILLAR_CODE, LV2 CHAPTER_CODE, LV3 SUB_CHAPTER_CODE, LV4 ITEM_CODE
                FROM (
                    --Pillar
                    SELECT S.YEAR, S.MONTH, S.FACTORY_CODE, P.ORDER_NO PILLAR_NO, -1 LEVEL_NO, 'Pillar' TYPE, P.PILLAR_CODE CODE,
                        CASE WHEN {is_eng} = 1 AND {is_vie} = 1 THEN CONCAT(P.PILLAR_ENG_NAME, ' / ', P.PILLAR_VIE_NAME)
                            WHEN {is_eng} = 1 AND {is_vie} = 0 THEN P.PILLAR_ENG_NAME
                            WHEN {is_eng} = 0 AND {is_vie} = 1 THEN P.PILLAR_VIE_NAME
                            ELSE '' END AS NAME,
                        CASE WHEN P.PILLAR_CODE <> 'SAF' THEN SUM(CASE WHEN S.SCORE <> -1 THEN S.SCORE END) END AS TOTAL_SCORE,
                        CASE WHEN P.PILLAR_CODE <> 'SAF' THEN COUNT(1) END ACTION,
                        CASE WHEN P.PILLAR_CODE <> 'SAF' THEN COUNT(CASE WHEN S.SCORE = -1 THEN 1 END) END ACTION_NA,
                        CASE WHEN P.PILLAR_CODE <> 'SAF' THEN ROUND(SUM(CASE WHEN S.SCORE <> -1 THEN S.SCORE END) / (COUNT(CASE WHEN S.SCORE <> -1 THEN 1 END) * 3) * 100, 2) END AS Assessment,
                        CASE WHEN P.PILLAR_CODE = 'ENV' THEN COUNT(CASE WHEN S.SCORE <> -1 THEN 1 END) * 3 ELSE NULL END Total_score_pillar,
                        CASE WHEN P.PILLAR_CODE = 'ENV' THEN SUM(CASE WHEN S.SCORE <> -1 THEN S.SCORE END) ELSE NULL END Final_score,
                        CASE WHEN P.PILLAR_CODE = 'ENV' THEN CAST(ROUND(SUM(CASE WHEN S.SCORE <> -1 THEN S.SCORE END) / (COUNT(CASE WHEN S.SCORE <> -1 THEN 1 END) * 3) * 100, 0) AS STRING) || '%' ELSE NULL END Grade,
                        NULL REDUCT_SCORING, NULL COMPLIANCE,
                        NULL CREATED_BY, NULL CREATED_DATE,
                        NULL LV4, NULL LV3, NULL LV2, P.PILLAR_CODE LV1
                    FROM udp_wcm_dev.MFG_OMS.OMS_SCORE_F S
                    LEFT JOIN udp_wcm_dev.MFG_OMS.OMS_TYPE_D KPI ON S.ITEM_CODE = KPI.CODE AND KPI.TYPE = 'Item'
                    LEFT JOIN udp_wcm_dev.MFG_OMS.OMS_TYPE_D SC ON KPI.PARENT_CODE = SC.CODE AND SC.TYPE = 'Sub_chapter'
                    LEFT JOIN udp_wcm_dev.MFG_OMS.OMS_TYPE_D C ON SC.PARENT_CODE = C.CODE AND C.TYPE = 'Chapter'
                    LEFT JOIN udp_wcm_dev.MFG_OMS.OMS_PILLAR_D P ON P.PILLAR_CODE = C.PILLAR_CODE
                    LEFT JOIN udp_wcm_dev.MFG_OMS.OMS_LEVEL_TYPE_D LV ON LV.TYPE = C.TYPE
                    WHERE S.YEAR = {year} AND S.MONTH = {month} AND S.FACTORY_CODE = '{factory_code}' AND P.PILLAR_CODE = '{pillar_code}'
                    GROUP BY S.YEAR, S.MONTH, S.FACTORY_CODE, P.ORDER_NO, P.PILLAR_CODE, P.PILLAR_ENG_NAME, P.PILLAR_VIE_NAME
                    UNION ALL --Chapter
                    SELECT YEAR, MONTH, FACTORY_CODE, PILLAR_NO, LEVEL_NO, TYPE, CODE,
                        NAME, TOTAL_SCORE,
                        ACTION, ACTION_NA, Assessment,
                        NULL Total_score_pillar, NULL Final_score, NULL Grade,
                        NULL REDUCT_SCORING, NULL COMPLIANCE,
                        NULL CREATED_BY, NULL CREATED_DATE,
                        NULL LV4, NULL LV3,
                        C.CODE LV2, C.PILLAR_CODE LV1
                    FROM C
                    UNION ALL --Chapter SAF00
                    SELECT YEAR, MONTH, FACTORY_CODE, PILLAR_NO, LEVEL_NO, TYPE, CODE,
                        NAME, SUM(TOTAL_SCORE) TOTAL_SCORE,
                        SUM(ACTION) ACTION, SUM(ACTION_NA) ACTION_NA, SUM(Assessment) Assessment,
                        NULL AS Total_score_pillar,
                        SUM(TOTAL_SCORE) - SUM(REDUCT_SCORING) Final_score,
                        CAST(ROUND((SUM(TOTAL_SCORE) - SUM(REDUCT_SCORING)) / (SUM(Assessment) * 100 * 3) * 100, 0) AS STRING) || '%' AS Grade,
                        ROUND(SUM(REDUCT_SCORING), 2) REDUCT_SCORING, C.COMPLIANCE,
                        NULL CREATED_BY, NULL CREATED_DATE,
                        NULL LV4, NULL LV3,
                        C.CODE LV2, C.PILLAR_CODE LV1
                    FROM C_SAF00 C
                    GROUP BY YEAR, MONTH, FACTORY_CODE, PILLAR_NO, LEVEL_NO, TYPE, CODE,
                        NAME,
                        C.CODE, C.PILLAR_CODE, C.COMPLIANCE
                    UNION ALL --Sub_chapter
                    SELECT YEAR, MONTH, FACTORY_CODE, PILLAR_NO, LEVEL_NO, TYPE, CODE,
                        NAME, TOTAL_SCORE,
                        ACTION, ACTION_NA, Assessment,
                        NULL Total_score_pillar, NULL Final_score, NULL Grade,
                        REDUCT_SCORING, COMPLIANCE,
                        NULL CREATED_BY, NULL CREATED_DATE,
                        NULL LV4, SC.CODE LV3,
                        SC.CHAPTER_CODE LV2, SC.PILLAR_CODE LV1
                    FROM SC
                    UNION ALL --KPI
                    SELECT S.YEAR, S.MONTH, S.FACTORY_CODE, P.ORDER_NO, LV.LEVEL_NO, KPI.TYPE, KPI.CODE,
                        CASE WHEN {is_eng} = 1 AND {is_vie} = 1 THEN CONCAT(KPI.ENG_NAME, ' / ', KPI.VIE_NAME)
                            WHEN {is_eng} = 1 AND {is_vie} = 0 THEN KPI.ENG_NAME
                            WHEN {is_eng} = 0 AND {is_vie} = 1 THEN KPI.VIE_NAME
                            ELSE '' END AS NAME,
                        CASE WHEN S.SCORE <> -1 THEN S.SCORE ELSE -1 END AS TOTAL_SCORE,
                        CASE WHEN SCORE <> -1 THEN 1 ELSE 0 END ACTION,
                        CASE WHEN SCORE = -1 THEN 1 ELSE 0 END ACTION_NA,
                        NULL AS Assessment,
                        NULL Total_score_pillar, NULL Final_score, NULL Grade,
                        NULL REDUCT_SCORING, NULL COMPLIANCE,
                        S.CREATED_BY, S.CREATED_DATE,
                        KPI.CODE LV4, SC.CODE LV3, C.CODE LV2, C.PILLAR_CODE LV1
                    FROM udp_wcm_dev.MFG_OMS.OMS_SCORE_F S
                    LEFT JOIN udp_wcm_dev.MFG_OMS.OMS_TYPE_D KPI ON S.ITEM_CODE = KPI.CODE AND KPI.TYPE = 'Item'
                    LEFT JOIN udp_wcm_dev.MFG_OMS.OMS_TYPE_D SC ON KPI.PARENT_CODE = SC.CODE AND SC.TYPE = 'Sub_chapter'
                    LEFT JOIN udp_wcm_dev.MFG_OMS.OMS_TYPE_D C ON SC.PARENT_CODE = C.CODE AND C.TYPE = 'Chapter'
                    LEFT JOIN udp_wcm_dev.MFG_OMS.OMS_PILLAR_D P ON P.PILLAR_CODE = C.PILLAR_CODE
                    LEFT JOIN udp_wcm_dev.MFG_OMS.OMS_LEVEL_TYPE_D LV ON LV.TYPE = KPI.TYPE
                    WHERE S.YEAR = {year} AND S.MONTH = {month} AND S.FACTORY_CODE = '{factory_code}' AND P.PILLAR_CODE = '{pillar_code}'
                )
                ORDER BY PILLAR_NO, CODE
            """
                
                cursor.execute(query)
                results = cursor.fetchall()

        # Convert results to DataFrame
        # return func.HttpResponse(json.dumps(results)) #return 186 rows - validated
                columns = [desc[0] for desc in cursor.description]
        # return func.HttpResponse(json.dumps(len(columns))) # "error": "'NoneType' object is not iterable"
        rows = [dict(zip(columns, row)) for row in results]
        # rows = [dict(row) for row in results]
        # return func.HttpResponse(json.dumps(rows))
        df = pd.DataFrame(rows)


        # Xử lý file Excel
        if file_type == "excel":
            for col in df.select_dtypes(include=['datetime64[ns, UTC]']).columns:
                df[col] = df[col].dt.tz_localize(None)

            output = BytesIO()
            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                df.to_excel(writer, index=False, sheet_name="TEMPLATE")
                worksheet = writer.sheets["TEMPLATE"]

                # Iterate through the DataFrame and format both columns and rows
                for col_idx, column in enumerate(df.columns, start=1):
                    col_letter = get_column_letter(col_idx)
                    
                    # Set column width and wrap_text for column "NAME"
                    if column == "NAME":
                        worksheet.column_dimensions[col_letter].width = 25
                        wrap_alignment = Alignment(wrap_text=True, vertical="top")
                    else:
                        column_width = max(df[column].astype(str).map(len).max(), len(column)) + 2
                        worksheet.column_dimensions[col_letter].width = column_width
                        wrap_alignment = Alignment(vertical="top")

                    # Format each cell in the column
                    for row_idx, value in enumerate(df[column], start=2):  # Start at row 2 to skip header
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        cell.value = value
                        cell.alignment = wrap_alignment

                        # Determine the format based on LEVEL_NO
                        level_no = df.at[row_idx - 2, "LEVEL_NO"] if "LEVEL_NO" in df.columns else None  # Adjust index for header

                        if int(level_no or 0) == 0: #Chapter
                            cell.fill = PatternFill(start_color="fff5ce", end_color="fff5ce", fill_type="solid")
                            cell.font = Font(color="000000", bold=True)

                        elif int(level_no or 0) == -1: #Pillar
                            cell.fill = PatternFill(start_color="E5F1FF", end_color="E5F1FF", fill_type="solid")
                            cell.font = Font(color="000000", bold=True)

                        elif int(level_no or 0) == 1: #Sub_Chapter
                            cell.fill = PatternFill(start_color="f6f6f3", end_color="f6f6f3", fill_type="solid")
                            cell.font = Font(color="000000", bold=True)

                        else:
                            cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                            cell.font = Font(color="000000", italic=True)

            output.seek(0)
            file_data = output.getvalue()
            file_extension = "xlsx"
            content_type = "vnd.openxmlformats-officedocument.spreadsheetml.sheet"

        else:  # CSV
            output = StringIO()
            df.to_csv(output, index=False, encoding="utf-8-sig")
            file_data = output.getvalue()
            file_extension = "csv"
            content_type = "text/csv"

        # Gửi email với file đính kèm
        filename = f"OMS_Result_{year}_{month}_{factory_code}_{pillar_code}.{file_extension}"
        send_email(file_data, filename, email_to, content_type)

        return func.HttpResponse(
            json.dumps({"message": "Xuất Kết quả OMS thành công, vui lòng kiểm tra Email để lấy file.", "code": 200}, ensure_ascii=False, indent=4), 
            status_code=200, 
            headers=headers
        )
        
    except jwt.ExpiredSignatureError:
        return func.HttpResponse(
            json.dumps({'message':"Token đã hết hạn!"}), 
            status_code=401, 
            headers=headers
        )
    except socket.timeout:
        return func.HttpResponse(
            json.dumps({"code": 500, "status": "error", "message": "Connection timeout - check managed identity configuration"}), 
            status_code=500, 
            headers=headers
        )
    except urllib.error.URLError as e:
        return func.HttpResponse(
            json.dumps({"code": 500, "status": "error", "message": f"URL Error: {str(e)}"}), 
            status_code=500, 
            headers=headers
        )
    except Exception as e:
        return func.HttpResponse(
            json.dumps({"error": str(e)}, ensure_ascii=False, indent=4), 
            status_code=500, 
            headers=headers
        )

def send_email(file_data, filename, email_to, content_type):
    """Gửi email với file Excel hoặc CSV đính kèm."""
    try:
        # Tạo email
        msg = MIMEMultipart()
        msg["From"] = email_from
        msg["To"] = email_to
        msg["Subject"] = "[OMS] - Result"

        body = "Kính gửi Anh/Chị,\n\nKết quả đã được xuất trong file đính kèm.\n\nTrân trọng,\n"
        msg.attach(MIMEText(body, 'plain'))

        # Đính kèm file
        attachment = MIMEBase("application", content_type)
        attachment.set_payload(file_data)
        encoders.encode_base64(attachment)
        attachment.add_header("Content-Disposition", f"attachment; filename={filename}")
        msg.attach(attachment)

        # Kết nối và gửi email
        server = smtplib.SMTP_SSL(host, port)
        server.login(email_from, gmail_password)
        server.sendmail(email_from, email_to, msg.as_string())
        server.close()
    except Exception as e:
        raise Exception(f"Lỗi gửi email: {str(e)}")

def decrypt_aes_json(encrypted_text):
    """
    Decrypt AES encrypted text and return JSON object.
    """
    try:
        if not encrypted_text or len(encrypted_text.strip()) == 0:
            raise ValueError("Dữ liệu mã hóa rỗng!")

        print("Encrypted Input (Base64):", encrypted_text)

        # Giải mã Base64
        encrypted_bytes = base64.b64decode(encrypted_text)

        print("Encrypted Bytes Length:", len(encrypted_bytes))

        # Kiểm tra độ dài dữ liệu hợp lệ
        if len(encrypted_bytes) % 16 != 0:
            raise ValueError("Dữ liệu mã hóa không phải bội số của 16 bytes!")

        # Giải mã AES
        cipher = Cipher(algorithms.AES(SECRET_KEY), modes.CBC(IV), backend=default_backend())
        decryptor = cipher.decryptor()
        decrypted_padded = decryptor.update(encrypted_bytes) + decryptor.finalize()

        print("Decrypted (Raw with Padding):", decrypted_padded)

        # Gỡ padding
        unpadder = padding.PKCS7(128).unpadder()
        decrypted_text = unpadder.update(decrypted_padded) + unpadder.finalize()

        print("Decrypted Text:", decrypted_text.decode())

        # Chuyển thành JSON
        return json.loads(decrypted_text.decode())
    
    except json.JSONDecodeError as je:
        print(f"JSONDecodeError: {je}")
    except ValueError as ve:
        print(f"ValueError: {ve}")
    except Exception as e:
        print(f"Decryption failed: {e}")
    
    return None  # Trả về None nếu có lỗi